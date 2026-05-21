# appointments/exports.py
import logging
from io import BytesIO

import openpyxl
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from reportlab.lib import colors
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.styles import ParagraphStyle, getSampleStyleSheet
from reportlab.lib.units import cm
from reportlab.platypus import Paragraph, SimpleDocTemplate, Spacer, Table, TableStyle

from django.http import HttpResponse
from django.utils import timezone

logger = logging.getLogger(__name__)


def export_citas_excel(queryset):
    """Returns an HttpResponse with an .xlsx attachment for the given queryset."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = 'Reporte de Citas'

    headers = [
        '#', 'Paciente', 'Email Paciente', 'Médico',
        'Especialidad', 'Fecha', 'Hora', 'Estado', 'Motivo',
    ]

    header_font = Font(bold=True, color='FFFFFF')
    header_fill = PatternFill(start_color='0d6efd', end_color='0d6efd', fill_type='solid')
    header_align = Alignment(horizontal='center', vertical='center')

    for col, text in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=text)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_align

    white_fill = PatternFill(start_color='FFFFFF', end_color='FFFFFF', fill_type='solid')
    gray_fill = PatternFill(start_color='f2f2f2', end_color='f2f2f2', fill_type='solid')

    for row_idx, cita in enumerate(queryset, 2):
        row_fill = gray_fill if row_idx % 2 == 0 else white_fill
        row_data = [
            cita.pk,
            cita.paciente.user.get_full_name(),
            cita.paciente.user.email,
            cita.medico.user.get_full_name(),
            str(cita.medico.especialidad),
            cita.fecha_hora.strftime('%d/%m/%Y'),
            cita.fecha_hora.strftime('%H:%M'),
            cita.get_estado_display(),
            cita.motivo,
        ]
        for col, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col, value=value)
            cell.fill = row_fill

    # Auto-adjust column widths
    for col in ws.columns:
        col_letter = get_column_letter(col[0].column)
        max_length = max(
            (len(str(cell.value)) for cell in col if cell.value is not None),
            default=0,
        )
        ws.column_dimensions[col_letter].width = max_length + 4

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    fecha = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(
        buffer.getvalue(),
        content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
    )
    response['Content-Disposition'] = f'attachment; filename="citas_{fecha}.xlsx"'
    return response


def export_citas_pdf(queryset, filtros_texto=''):
    """Returns an HttpResponse with a .pdf attachment for the given queryset."""
    buffer = BytesIO()
    doc = SimpleDocTemplate(
        buffer,
        pagesize=landscape(letter),
        leftMargin=1.5 * cm,
        rightMargin=1.5 * cm,
        topMargin=1.5 * cm,
        bottomMargin=1.5 * cm,
    )

    styles = getSampleStyleSheet()
    story = []

    # Title
    title_style = ParagraphStyle(
        'ReportTitle',
        parent=styles['Heading1'],
        alignment=1,
        spaceAfter=6,
    )
    story.append(Paragraph('Reporte de Citas Médicas', title_style))

    # Subtitle with active filters
    if filtros_texto:
        subtitle_style = ParagraphStyle(
            'FilterSubtitle',
            parent=styles['Normal'],
            fontSize=9,
            alignment=1,
            textColor=colors.grey,
            spaceAfter=4,
        )
        story.append(Paragraph(f'Filtros: {filtros_texto}', subtitle_style))

    # Generation date (right-aligned)
    date_style = ParagraphStyle(
        'GenDate',
        parent=styles['Normal'],
        fontSize=8,
        alignment=2,
        textColor=colors.grey,
        spaceAfter=12,
    )
    story.append(
        Paragraph(f"Generado: {timezone.now().strftime('%d/%m/%Y %H:%M')}", date_style)
    )

    rows = list(queryset)

    if not rows:
        story.append(
            Paragraph('No hay citas con los filtros aplicados.', styles['Normal'])
        )
    else:
        col_headers = ['Paciente', 'Médico', 'Especialidad', 'Fecha/Hora', 'Estado']
        data = [col_headers]

        for cita in rows:
            data.append([
                cita.paciente.user.get_full_name(),
                cita.medico.user.get_full_name(),
                str(cita.medico.especialidad),
                cita.fecha_hora.strftime('%d/%m/%Y %H:%M'),
                cita.get_estado_display(),
            ])

        table = Table(data, repeatRows=1, hAlign='LEFT')

        header_bg = colors.HexColor('#0d6efd')
        even_bg = colors.HexColor('#f8f9fa')
        grid_color = colors.HexColor('#dee2e6')

        ts = TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), header_bg),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.white),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('ALIGN', (0, 0), (-1, 0), 'CENTER'),
            ('GRID', (0, 0), (-1, -1), 0.5, grid_color),
            ('TOPPADDING', (0, 0), (-1, -1), 6),
            ('BOTTOMPADDING', (0, 0), (-1, -1), 6),
            ('LEFTPADDING', (0, 0), (-1, -1), 6),
            ('RIGHTPADDING', (0, 0), (-1, -1), 6),
            ('VALIGN', (0, 0), (-1, -1), 'MIDDLE'),
        ])
        # Alternate body row colors
        for i in range(2, len(data), 2):
            ts.add('BACKGROUND', (0, i), (-1, i), even_bg)

        table.setStyle(ts)
        story.append(table)

    doc.build(story)
    buffer.seek(0)

    fecha = timezone.now().strftime('%Y%m%d')
    response = HttpResponse(buffer.getvalue(), content_type='application/pdf')
    response['Content-Disposition'] = f'attachment; filename="citas_{fecha}.pdf"'
    return response
